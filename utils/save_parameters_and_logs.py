
#save_parameters_and_logs.py
# Experiment tracking is now handled by Weights & Biases (wandb) integration in train/mainLoop.py.
# The Excel and static image-based logging below is deprecated and retained for reference only.

# import os
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
# from openpyxl.chart import BarChart, Reference

# def save_params_and_logs(params, log_data, task_Assignments_info):
#     current_dir = os.path.dirname(os.path.abspath(__file__))

#     results_dir = os.path.join(current_dir, f'{params.SCENARIO_TYPE}_results')
#     if not os.path.exists(results_dir):
#         os.makedirs(results_dir)
    
#     filename = os.path.join(results_dir, f"Permutation_{params.Permutation_Number}.xlsx")

#     simulator_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
#     fname = os.path.join(simulator_dir, 'homogeneous_server_info.xlsx') if params.SCENARIO_TYPE == 'homogeneous' else os.path.join(simulator_dir, 'heterogeneous_server_info.xlsx')
#     sheet_name = f'{params.SCENARIO_TYPE.capitalize()}_Permutation_{params.Permutation_Number}'
#     server_info = pd.read_excel(fname, sheet_name=sheet_name)

#     task_df = pd.read_excel(os.path.join(simulator_dir, 'task_parameters.xlsx'))
#     task_info = task_df.values.tolist()

#     params_data = {attr: [value] for attr, value in vars(params).items()}
#     log_data = [{'Episode': log[0], 'Avg Reward': log[1], 'Episode Reward': log[2], 'Avg Delay': log[3]} for log in log_data]

#     df_params = pd.DataFrame(params_data).transpose().reset_index()
#     df_params.columns = ['Parameter', 'Value']
#     df_tasks = pd.DataFrame(task_info, columns=['Task_ID', 'Task_Size', 'Computation_Demand'])
#     server_columns = server_info.columns.tolist()
#     df_servers = pd.DataFrame(server_info, columns=server_columns)
#     df_logs = pd.DataFrame(log_data)
#     df_task_Assignments = pd.DataFrame(task_Assignments_info, columns=['episode', 'task_id', 'Primary', 'Primary_Start', 'Primary_End', 'Primary_Status', 'Backup', 'Backup_Start', 'Backup_End', 'Backup_Status', 'Z'])
#     df_task_Assignments['Final_status'] = df_task_Assignments.apply(
#         lambda row: 'failure' if row['Primary_Status'] == 'failure' and row['Backup_Status'] == 'failure' else 'success',
#         axis=1
#     )

#     summary_df = df_task_Assignments.groupby(['episode', 'Final_status']).size().unstack(fill_value=0)
#     summary_df = summary_df.rename(columns={'failure': 'Failure', 'success': 'Success'}).reset_index()

#     with pd.ExcelWriter(filename) as writer:
#         df_params.to_excel(writer, sheet_name='Params', index=False)
#         df_tasks.to_excel(writer, sheet_name='Tasks', index=False)
#         df_servers.to_excel(writer, sheet_name='Servers', index=False)
#         df_logs.to_excel(writer, sheet_name='Logs', index=False)
#         df_task_Assignments.to_excel(writer, sheet_name='TaskAssignments', index=False)
#         summary_df.to_excel(writer, sheet_name='Summary', index=False)

#     workbook = load_workbook(filename)
#     worksheet = workbook['TaskAssignments']
#     chart = BarChart()
#     chart.title = "Success and Failure Counts per Episode"
#     chart.x_axis.title = 'Episode'
#     chart.y_axis.title = 'Count'
    
#     data = Reference(workbook['Summary'], min_col=2, min_row=1, max_col=3, max_row=summary_df.shape[0] + 1)
#     categories = Reference(workbook['Summary'], min_col=1, min_row=2, max_row=summary_df.shape[0] + 1)
#     chart.add_data(data, titles_from_data=True)
#     chart.set_categories(categories)
    
#     worksheet.add_chart(chart, 'J5')  # Adjust the cell location as needed

#     worksheet_logs = workbook['Logs']
#     img_path = os.path.join(current_dir, 'rewards_plot.png')
#     img = Image(img_path)
#     worksheet_logs.add_image(img, 'E2')  # Adjust the cell location as needed

#     workbook.save(filename)
#     print("successfully saved logs in subfolders of our approach!")
    

'''
#save_parameters_and_logs
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt

def save_params_and_logs(params, log_data, task_Assignments_info):
    current_dir = os.path.dirname(os.path.abspath(__file__))

    results_dir = os.path.join(current_dir, f'{params.SCENARIO_TYPE}_results')
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
    
    filename = os.path.join(results_dir, f"Permutation_{params.Permutation_Number}.xlsx")
    hdf5_filename = os.path.join(results_dir, f"Permutation_{params.Permutation_Number}.h5")

    simulator_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
    fname = os.path.join(simulator_dir, 'homogeneous_server_info.xlsx') if params.SCENARIO_TYPE == 'homogeneous' else os.path.join(simulator_dir, 'heterogeneous_server_info.xlsx')
    sheet_name = f'{params.SCENARIO_TYPE.capitalize()}_Permutation_{params.Permutation_Number}'
    server_info = pd.read_excel(fname, sheet_name=sheet_name)

    task_df = pd.read_excel(os.path.join(simulator_dir, 'task_parameters.xlsx'))
    task_info = task_df.values.tolist()

    params_data = {attr: [value] for attr, value in vars(params).items()}
    log_data = [{'Episode': log[0], 'Avg Reward': log[1], 'Episode Reward': log[2]} for log in log_data]

    df_params = pd.DataFrame(params_data).transpose().reset_index()
    df_params.columns = ['Parameter', 'Value']
    df_tasks = pd.DataFrame(task_info, columns=['Task_ID', 'Task_Size', 'Computation_Demand'])
    server_columns = server_info.columns.tolist()
    df_servers = pd.DataFrame(server_info, columns=server_columns)
    df_logs = pd.DataFrame(log_data)
    df_task_Assignments = pd.DataFrame(task_Assignments_info, columns=['episode', 'task_id', 'Primary', 'Primary_Start', 'Primary_End', 'Primary_Status', 'Backup', 'Backup_Start', 'Backup_End', 'Backup_Status', 'Z'])
    df_task_Assignments['Final_status'] = df_task_Assignments.apply(
        lambda row: 1 if row['Primary_Status'] == 'failure' and row['Backup_Status'] == 'failure' else 0,
        axis=1
    )

    # Replace string values with numerical codes
    status_mapping = {'failure': 1, 'success': 0}
    df_task_Assignments['Primary_Status'] = df_task_Assignments['Primary_Status'].map(status_mapping)
    df_task_Assignments['Backup_Status'] = df_task_Assignments['Backup_Status'].map(status_mapping)

    summary_df = df_task_Assignments.groupby(['episode', 'Final_status']).size().unstack(fill_value=0)
    summary_df = summary_df.rename(columns={1: 'Failure', 0: 'Success'}).reset_index()

    # Save large data to HDF5
    with pd.HDFStore(hdf5_filename, mode='w') as store:
        store.put('task_assignments', df_task_Assignments, format='table')

    # Save smaller data to Excel
    with pd.ExcelWriter(filename) as writer:
        df_params.to_excel(writer, sheet_name='Params', index=False)
        df_tasks.to_excel(writer, sheet_name='Tasks', index=False)
        df_servers.to_excel(writer, sheet_name='Servers', index=False)
        df_logs.to_excel(writer, sheet_name='Logs', index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    workbook = load_workbook(filename)
    summary_sheet = workbook['Summary']
    chart = BarChart()
    chart.title = "Success and Failure Counts per Episode"
    chart.x_axis.title = 'Episode'
    chart.y_axis.title = 'Count'
    
    data = Reference(summary_sheet, min_col=2, min_row=1, max_col=3, max_row=summary_df.shape[0] + 1)
    categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=summary_df.shape[0] + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    summary_sheet.add_chart(chart, 'J5')  # Adjust the cell location as needed

    # Plot and save the rewards graph
    df_logs.plot(x='Episode', y=['Avg Reward', 'Episode Reward'], kind='line')
    plt.title('Rewards per Episode')
    plt.xlabel('Episode')
    plt.ylabel('Reward')
    img_path = os.path.join(current_dir, 'rewards_plot.png')
    plt.savefig(img_path)
    plt.close()

    worksheet_logs = workbook['Logs']
    img = Image(img_path)
    worksheet_logs.add_image(img, 'E2')  # Adjust the cell location as needed

    workbook.save(filename)
    print("Successfully saved logs in subfolders of our approach!")
'''